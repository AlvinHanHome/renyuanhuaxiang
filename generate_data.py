# -*- coding: utf-8 -*-
"""
生成「发展党员人员画像」演示数据 + 导入模板。
输出：
  demo-data.js      —— 前端演示数据 (window.DEMO_PEOPLE)
  sample-data.xlsx  —— 可直接导入的 Excel 模板（含填写说明）

积分规则：每维满分20分，从0分开始累积；总分=五维之和（无加权）；按总分排序择优。

单位与党支部严格参照《党支部信息表-模板.xlsx》：
  11 个单位（10 个县区局 + 周口市烟草公司卷烟物流配送中心），共 83 个党支部。
"""
import random, json, datetime, os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

random.seed(20260822)

BASE = os.path.dirname(os.path.abspath(__file__))


CN = {1:"一",2:"二",3:"三",4:"四",5:"五",6:"六",7:"七",8:"八",9:"九",10:"十"}
def numbered(prefix, start, end, skip=None):
    """生成 '前缀+中文数字+党支部'，可跳过某些序号（如沈丘缺第三党支部）。"""
    skip = skip or set()
    return [f"{prefix}第{CN[i]}党支部" for i in range(start, end + 1) if i not in skip]


# ---------- 11 个单位 + 对应党支部（与党支部信息表-模板.xlsx 完全一致） ----------
COUNTIES = [
    ("周口市川汇区烟草专卖局(分公司)", numbered("周口市川汇区烟草专卖局", 1, 10)),
    ("鹿邑县烟草专卖局(分公司)",       numbered("鹿邑县烟草专卖局", 1, 7)),
    ("商水县烟草专卖局(分公司)",       numbered("商水县烟草专卖局", 1, 8)),
    ("郸城县烟草专卖局(分公司)",       numbered("郸城县烟草专卖局", 1, 8)),
    ("周口市淮阳区烟草专卖局(分公司)", numbered("周口市淮阳区烟草专卖局", 1, 9)),
    ("项城市烟草专卖局(分公司)",       numbered("项城市烟草专卖局", 1, 10)),
    # 模板中太康县写作"太康县烟草局"（非"烟草专卖局"）
    ("太康县烟草专卖局(分公司)",       numbered("太康县烟草局", 1, 7)),
    # 模板中沈丘县缺"第三党支部"
    ("沈丘县烟草专卖局(分公司)",       numbered("沈丘县烟草专卖局", 1, 6, skip={3})),
    ("西华县烟草专卖局(分公司)",       numbered("西华县烟草专卖局", 1, 5)),
    ("扶沟县烟草专卖局(分公司)",       numbered("扶沟县烟草专卖局", 1, 6)),
    ("周口市烟草公司卷烟物流配送中心", [
        "周口市卷烟物流配送中心技术部党支部",
        "周口市卷烟物流配送中心鹿邑中转站党支部",
        "周口市卷烟物流配送中心太康中转站党支部",
        "周口市卷烟物流配送中心综合部党支部",
        "周口市卷烟物流配送中心送货部党支部",
        "周口市卷烟物流配送中心安保部党支部",
        "周口市卷烟物流配送中心财务部党支部",
        "周口市卷烟物流配送中心储配部党支部",
    ]),
]

STAGES = ["申请入党", "入党积极分子", "发展对象", "预备党员"]
STAGE_W = [0.34, 0.34, 0.20, 0.12]

# 岗位属性（与积分制管理细则一致）
JOB_TYPES = ["综合岗", "专卖岗", "销售岗", "配送岗"]
JOB_W = [0.22, 0.28, 0.30, 0.20]

# ---------- 姓名池 ----------
SURNAMES = list("王李张刘陈杨赵黄周吴徐孙马朱胡郭何高林郑谢罗梁宋唐许韩冯邓曹彭曾肖田董袁潘于蒋蔡余杜叶程苏魏吕丁任沈姚卢傅钟姜崔谭廖范汪廖")
GIVEN_M = ["伟","强","磊","军","勇","杰","涛","斌","波","辉","鹏","宇","浩然","建国","志远","晓东","文博","嘉豪","宇航","泽宇","瑞","凯","鑫","俊","晨"]
GIVEN_F = ["芳","娟","敏","静","丽","艳","玲","燕","娜","秀英","霞","倩","雪","婷","琳","佳","悦","欣怡","梦洁","雨彤","思琪","雅楠","诗涵","梓涵","欣","璐"]

def id_card(county_idx):
    region = "4116" + f"{county_idx:02d}"
    y = random.randint(1992, 2002)
    m = random.randint(1, 12); d = random.randint(1, 28)
    birth = f"{y:04d}{m:02d}{d:02d}"
    seq = f"{random.randint(0,999):03d}"
    body = region + birth + seq
    w = [7,9,10,5,8,4,2,1,6,3,7,9,10,5,8,4,2]
    codes = "10X98765432"
    s = sum(int(body[i])*w[i] for i in range(17))
    body += codes[s % 11]
    return body

def phone():
    return "1" + random.choice("356789") + "".join(random.choice("0123456789") for _ in range(9))

def apply_date():
    start = datetime.date(2022, 1, 1)
    end = datetime.date(2025, 6, 30)
    d = start + datetime.timedelta(days=random.randint(0, (end-start).days))
    return d.strftime("%Y-%m-%d")

def scores():
    """每维 0~20 分，从 0 开始累积；总分 ≤ 100。"""
    return {
        "思想态度": random.randint(8, 20),
        "学习成效": random.randint(7, 20),
        "日常表现": random.randint(9, 20),
        "履职担当": random.randint(8, 20),
        "工作实绩": random.randint(10, 20),
    }

people = []
pid = 0
# 各单位人数分布（合计 64 人，覆盖全部 11 个单位）
counts = [8, 6, 6, 6, 7, 7, 5, 4, 4, 5, 6]
for ci, (county, orgs) in enumerate(COUNTIES):
    for _ in range(counts[ci]):
        pid += 1
        gender = random.choice(["男", "女"])
        if gender == "男":
            name = random.choice(SURNAMES) + random.choice(GIVEN_M)
        else:
            name = random.choice(SURNAMES) + random.choice(GIVEN_F)
        stage = random.choices(STAGES, weights=STAGE_W, k=1)[0]
        job = random.choices(JOB_TYPES, weights=JOB_W, k=1)[0]
        org = random.choice(orgs)            # 真实党支部名称
        sc = scores()
        people.append({
            "name": name,
            "gender": gender,
            "idCard": id_card(ci),
            "phone": phone(),
            "county": county,                # 县（区）局：全称
            "org": "中共" + org,             # 申请受理党组织
            "applyDate": apply_date(),
            "stage": stage,
            "jobType": job,
            "scores": sc,
        })

# 设置 2 名标杆人员（便于对比展示）
people[0].update({"name":"韩继伟","gender":"男","stage":"入党积极分子","jobType":"综合岗",
                  "scores":{"思想态度":19,"学习成效":18,"日常表现":17,"履职担当":18,"工作实绩":20},
                  "county":"鹿邑县烟草专卖局(分公司)","org":"中共鹿邑县烟草专卖局第一党支部","applyDate":"2023-03-12"})
people[1].update({"name":"苏瑞","gender":"男","stage":"发展对象","jobType":"专卖岗",
                  "scores":{"思想态度":17,"学习成效":19,"日常表现":18,"履职担当":16,"工作实绩":19},
                  "county":"扶沟县烟草专卖局(分公司)","org":"中共扶沟县烟草专卖局第一党支部","applyDate":"2024-09-27"})

# 修正标杆身份证/电话
people[0]["idCard"] = id_card(1); people[0]["phone"] = phone()
people[1]["idCard"] = id_card(9); people[1]["phone"] = phone()

# 按总分排序
def total(p): return sum(p["scores"].values())
people.sort(key=total, reverse=True)

# ---------- 写出 demo-data.js ----------
js = "// 演示数据：入党申请人基本信息与五维积分（由 generate_data.py 自动生成）\n"
js += "window.DEMO_PEOPLE = " + json.dumps(people, ensure_ascii=False, indent=2) + ";\n"
with open(os.path.join(BASE, "demo-data.js"), "w", encoding="utf-8") as f:
    f.write(js)

# ---------- 写出 sample-data.xlsx ----------
HEAD = ["姓名","性别","身份证号","联系电话","县（区）局","申请受理党组织","申请书递交日期","当前阶段","岗位属性","思想态度","学习成效","日常表现","履职担当","工作实绩"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "入党申请人信息"

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="C0182C")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
center = Alignment(horizontal="center", vertical="center")

ws.append(HEAD)
for c in range(1, len(HEAD)+1):
    cell = ws.cell(row=1, column=c)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border

for p in people:
    s = p["scores"]
    row = [p["name"], p["gender"], p["idCard"], p["phone"], p["county"], p["org"],
           p["applyDate"], p["stage"], p["jobType"],
           s["思想态度"], s["学习成效"], s["日常表现"], s["履职担当"], s["工作实绩"]]
    ws.append(row)
    for c in range(1, len(HEAD)+1):
        ws.cell(row=ws.max_row, column=c).border = border
        ws.cell(row=ws.max_row, column=c).alignment = Alignment(horizontal="center", vertical="center")

widths = [10,6,20,14,30,32,14,12,10,9,9,9,9,9]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# 填写说明
note = wb.create_sheet("填写说明")
note_lines = [
    ["发展党员人员画像 · 导入模板填写说明"],
    [""],
    ["1. 本模板核心工作表为【入党申请人信息】，请勿修改表头名称，行顺序不限。"],
    ["2. 基本信息字段：姓名、性别、身份证号、联系电话、县（区）局、申请受理党组织、申请书递交日期。"],
    ["3. 当前阶段：填写 申请入党 / 入党积极分子 / 发展对象 / 预备党员。"],
    ["4. 岗位属性：填写 综合岗 / 专卖岗 / 销售岗 / 配送岗。"],
    ["5. 五维积分（核心）：思想态度、学习成效、日常表现、履职担当、工作实绩。"],
    ["   · 每个维度满分 20 分，从 0 分开始累积；"],
    ["   · 总分 = 五个维度得分之和（满分 100 分），按总分排名择优发展。"],
    ["6. 县（区）局、申请受理党组织须与《党支部信息表》中的规范名称一致。"],
    ["7. 填写完成后，在页面右上角【数据导入】面板选择本文件即可载入。"],
    [""],
    ["生成时间：" + datetime.datetime.now().strftime("%Y-%m-%d %H:%M")],
]
for ln in note_lines:
    note.append(ln)
note["A1"].font = Font(bold=True, size=14, color="C0182C")

wb.save(os.path.join(BASE, "sample-data.xlsx"))
print(f"生成完成：{len(people)} 人（{len(COUNTIES)} 个单位 / {sum(len(o) for _,o in COUNTIES)} 个党支部）；demo-data.js + sample-data.xlsx 已写出。")
print("样例前3：", [(p["name"], p["county"], p["org"], sum(p["scores"].values())) for p in people[:3]])
